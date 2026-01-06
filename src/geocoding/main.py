#!/usr/bin/env python

"""
Geocoding CLI Tool（百度地图版）：将 Excel 中指定列的地址转换为经纬度
用法示例：
    python geocoding.py --key your_baidu_ak --sheet 1 --column 5 input.xlsx
"""

import argparse
import os
import re
import sys
import time
from pathlib import Path

import openpyxl
import requests
from tqdm import tqdm

BAIDU_GEOCODE_URL = "https://api.map.baidu.com/geocoding/v3/"


def geocode_address(address: str, key: str) -> (float, float):
    """使用百度地图 API 将地址转为经纬度（BD09 坐标系）"""
    if not address or not str(address).strip():
        return None, None

    params = {
        "address": str(address).strip(),
        "output": "json",
        "ak": key,
        # 可选参数：城市限定（提升准确率），此处暂不使用
        # "city": "全国"
        'ret_coordtype': 'gcj02ll'
    }
    try:
        response = requests.get(BAIDU_GEOCODE_URL, params=params, timeout=10)
        data = response.json()
        # 百度 API 成功时 status 为 0
        if data.get("status") == 0:
            result = data["result"]
            location = result["location"]
            lng = location["lng"]
            lat = location["lat"]
            return float(lng), float(lat)
        else:
            # 例如：{"status":1,"message":"Invalid AK"}
            print(f"Geocoding failed for '{address}': {data.get('message', 'Unknown error')}", file=sys.stderr)
            return None, None
    except Exception as e:
        print(f"Request error for '{address}': {e}", file=sys.stderr)
        return None, None


def main():
    parser = argparse.ArgumentParser(
        description="将 Excel 文件中指定列的地址转换为经纬度（百度地图 API）"
    )
    parser.add_argument("input_file", help="输入的 .xlsx 文件路径")
    parser.add_argument("--key", "-k", required=True, help="百度地图 Web 服务 AK（Access Key）")
    parser.add_argument("--sheet", "-s", type=int, default=1, help="工作表序号（从1开始，默认为1）")
    parser.add_argument("--column", "-c", type=int, required=True, help="地址所在列号（从1开始，A=1, B=2...）")
    parser.add_argument('--sleep', '-sl', type=float, default=0.2, help="请求间隔时间（秒）")

    args = parser.parse_args()

    input_file = args.input_file
    key = args.key
    sheet_index = args.sheet
    col_index = args.column
    _间隔时间 = args.sleep

    if not os.path.isfile(input_file):
        print(f"错误：文件 '{input_file}' 不存在。", file=sys.stderr)
        sys.exit(1)

    if not input_file.lower().endswith('.xlsx'):
        print("错误：仅支持 .xlsx 格式文件。", file=sys.stderr)
        sys.exit(1)

    try:
        _workbook = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"无法打开 Excel 文件: {e}", file=sys.stderr)
        sys.exit(1)

    if sheet_index < 1 or sheet_index > len(_workbook.sheetnames):
        print(f"错误：工作表序号 {sheet_index} 超出范围（共 {len(_workbook.sheetnames)} 个表）。", file=sys.stderr)
        sys.exit(1)

    _worksheet = _workbook.worksheets[sheet_index - 1]

    if col_index < 1 or col_index > _worksheet.max_column:
        print(f"错误：列号 {col_index} 超出工作表列范围（最大列: {_worksheet.max_column}）。", file=sys.stderr)
        sys.exit(1)

    # 新增两列用于写入结果
    lng_col = _worksheet.max_column + 1
    lat_col = _worksheet.max_column + 2
    _worksheet.cell(row=1, column=lng_col, value="经度")
    _worksheet.cell(row=1, column=lat_col, value="纬度")

    print(f"正在处理工作表: {_worksheet.title}")
    print(f"读取地址列（第 {col_index} 列），输出到新列 {openpyxl.utils.get_column_letter(lng_col)} 和 {openpyxl.utils.get_column_letter(lat_col)}...")

    success_count = 0
    total_rows = _worksheet.max_row

    with tqdm(total=total_rows, desc=f"地理编码中...") as pbar:
        for row in range(2, total_rows + 1):  # 从第2行开始（跳过标题）
        # for row in range(2, 3):  # 从第2行开始（跳过标题）
            address_cell = _worksheet.cell(row=row, column=col_index)
            _地址 = address_cell.value
            print(f'正在处理 {_地址}')

            if re.match(r'^\*+$', _地址):
                print(f'{_地址} 是私密地址，跳过')
                continue

            lng, lat = geocode_address(_地址, key)
            print(f'经度: {lng}, 纬度: {lat}')
            if lng is not None and lat is not None:
                _worksheet.cell(row=row, column=lng_col, value=lng)
                _worksheet.cell(row=row, column=lat_col, value=lat)
                success_count += 1

            # 百度 API 免费配额通常为 10万次/天，QPS <= 30，建议延时 ≥0.5s
            time.sleep(_间隔时间)

            pbar.update(n=1)

    try:
        _path = Path(input_file)
        _新文件名 = _path.stem + '_已加入经纬度' + _path.suffix
        _workbook.save(f'{_path.parent / _新文件名}')
        print(f"\n✅ 完成！共处理 {total_rows - 1} 行，成功转换 {success_count} 个地址。")
        print(f"结果已写入文件: {_path.parent / _新文件名}")
        print("注意：经纬度为百度 BD09 坐标系，如需 WGS84 请另行转换。")
    except Exception as e:
        print(f"保存文件失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
